import openpyxl
import sys
import warnings
from openpyxl.styles import Font, Alignment
import win32com.client as win32
import os

n = len(sys.argv)
FOLDER_NAME = "gens"
PATH_TO_DIR = str(os.getcwd() + f"\\{FOLDER_NAME}\\")
if not os.path.isdir(PATH_TO_DIR):
    os.mkdir(PATH_TO_DIR)

# Set the year, month, and starting day (defaults):
YEAR = "2024"
MONTH = "AUGUST"
DAY = "THURSDAY"
TIME = [YEAR, MONTH, DAY]

# Validate n and print appropriate message
if n not in [1, 2, 4, 5]:
    raise ValueError("Invalid number of arguments. Please pass either 0 arguments (for default settings) or 3 arguments (for YEAR, MONTH, DAY).")
elif n in [1, 2]:
    print(f"Using default settings: YEAR={YEAR}, MONTH={MONTH}, DAY={DAY}")
elif n in [4, 5]:
    # If year, month, and starting day were passed as arguments, use those instead of default:
    for i in range(1,4):
        TIME[i-1] = str(sys.argv[i]).upper() 
    print(f"Using user input for {TIME[0].lower()}, {TIME[1].lower()}, and {TIME[2].lower()}.")

# Suppress specific warnings related to openpyxl header/footer parsing
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet.header_footer")

# Load the existing Excel workbook
workbook = openpyxl.load_workbook('kitchen_plan.xlsx')

# Select the active worksheet (or specify a sheet by name)
sheet = workbook.active  # or workbook['SheetName']

# For naming the output files
FILESUFFIX = ""
for i in range(len(TIME)):
    FILESUFFIX += f'_{TIME[i].lower()}'

# Set the year, month, and day in the excel template
sheet.cell(row=1, column=2).value = TIME[0]
sheet.cell(row=1, column=3).value = TIME[1]
sheet.cell(row=1, column=5).value = TIME[2]

# Flatmate names:

# If the user gave one extra command line argument and wants to specify flatmates
if n in [2, 5]:
    while True:
        try:
            # Read user input for names, split by spaces
            input_names = input("Type the names of the flatmates separated by spaces: ")
            NAMES = input_names.split()

            # Check if the list is empty or too short
            if len(NAMES) < 2:
                raise ValueError("Please enter at least two names.")

            # Print the list for confirmation
            print("\nYou entered the following names:")
            for name in NAMES:
                print(name)

            # Ask the user if the list is correct
            confirmation = input("\nDoes this look correct? (y/n): ").strip().lower()
            if confirmation == 'y':
                break
            else:
                print("Let's try again.\n")
        except ValueError as e:
            print(f"Error: {e}")
            print("Please try again.\n")

# If the user didn't give the extra argument and wants to use the hardcoded names
else:
    NAMES = ["Sol", "Georg", "Burak", "Ishaan", "Kenna", "Julia", "Sophie", "Paula"]
    print(f"Creating schedule using these flatmates:\n{NAMES}\nTo use a different list of flatmates, edit the python script or give one more command line argument (it can be anything) to provide the flatmate names directly in the command line (I'm not making a GUI)")

# Edit the cells in the columns B-H, rows 5, 7, 9, 11, 13, and 15
name_index = 0
for i in range(6):
    for j in range(7):
        cell = sheet.cell(row=5+i*2, column=2+j)
        cell.value = NAMES[name_index%len(NAMES)]
        cell.font = Font(size=20)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        name_index += 1

# Save the edited workbook to a new file or overwrite the existing file
workbook.save(f'{FOLDER_NAME}\\kitchen_plan{FILESUFFIX}.xlsx')

##### PRINT TO PDF #####

# Load the Excel application
excel = win32.Dispatch('Excel.Application')

# Open the Excel workbook
workbook = excel.Workbooks.Open(PATH_TO_DIR + f'kitchen_plan{FILESUFFIX}.xlsx')

# Specify the sheet you want to convert to PDF (0 for the active sheet, or specify by name)
sheet = workbook.Worksheets[0]

# Save the sheet as PDF
pdf_path = PATH_TO_DIR + f'kitchen_plan{FILESUFFIX}.pdf'
sheet.ExportAsFixedFormat(0, pdf_path)

# Close the workbook and Excel application
workbook.Close(False)
excel.Quit()

print(f"Excel sheet has been saved as a PDF file: {pdf_path}")
